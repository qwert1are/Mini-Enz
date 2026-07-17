"""
MiniEnz — Organize all results into a comprehensive Excel workbook with 6 sheets.
"""
import json
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, BarChart3D

OUT_PATH = r"H:\eazyclaw\saved\MiniEnz_Methodology\reports\MiniEnz_Results.xlsx"
ANALYSIS_DIR = r"H:\eazyclaw\saved\MiniEnz_Methodology\results\analysis"

wb = Workbook()

# --- Styles ---
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
subheader_fill = PatternFill(start_color='D2691E', end_color='D2691E', fill_type='solid')
best_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
bad_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt

def auto_width(ws):
    for col_idx in range(1, ws.max_column+1):
        max_len = 0
        for row_idx in range(1, ws.max_row+1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

# ============================================================
# SHEET 1: Summary Dashboard
# ============================================================
ws1 = wb.active
ws1.title = "Summary Dashboard"

# Title
ws1.merge_cells('A1:H1')
ws1['A1'] = 'MiniEnz — Enzyme Miniaturization Benchmark Summary'
ws1['A1'].font = Font(name='Calibri', bold=True, size=14, color='8B4513')
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:H2')
ws1['A2'] = 'June 10, 2026 | 8 Enzymes | 1,800 RFdiffusion Backbones | 4,800 ProteinMPNN Sequences'
ws1['A2'].font = Font(name='Calibri', size=10, color='666666')
ws1['A2'].alignment = Alignment(horizontal='center')

# KPIs
kpi_data = [
    ['Metric', 'Value', 'Unit'],
    ['Enzyme families', 8, ''],
    ['EC classes covered', 6, ''],
    ['Fold types', 7, ''],
    ['RFdiffusion backbones', 1800, 'PDB files'],
    ['ProteinMPNN sequences', 4800, 'FASTA entries'],
    ['ESM-2 embeddings', 4800, '640-dim vectors'],
    ['SRE range', '6.0 – 50.2', '%'],
    ['Catalytic RMSD < 1.0 Å', '100', '% of designs'],
    ['Independent dims (r≈0)', 'Catalytic RMSD ⊥ MPNN Score', ''],
    ['Best MPNN score', 0.862, 'TIM barrel'],
    ['Best SRE', 50.2, '% (pc_lipase, multi-motif)'],
]
for i, row in enumerate(kpi_data, start=4):
    for j, val in enumerate(row):
        ws1.cell(row=i, column=j+1, value=val)
style_header_row(ws1, 4, 3)
for i in range(5, 15):
    for j in range(1, 4):
        style_cell(ws1, i, j)

auto_width(ws1)
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 25

# ============================================================
# SHEET 2: Size Reduction Efficiency
# ============================================================
ws2 = wb.create_sheet("1-Size Reduction (SRE)")

ws2.merge_cells('A1:I1')
ws2['A1'] = 'Size Reduction Efficiency (SRE) — All 8 Enzymes'
ws2['A1'].font = Font(name='Calibri', bold=True, size=13, color='8B4513')

sre_headers = ['Enzyme', 'PDB', 'EC', 'Fold Type', 'Native (aa)', 'Designed Mean (aa)', 
               'Designed SD (aa)', 'SRE (%)', 'Strategy']
for j, h in enumerate(sre_headers, 1):
    ws2.cell(row=3, column=j, value=h)
style_header_row(ws2, 3, len(sre_headers))

sre_data = [
    ['P. cepacia Lipase', '3LIP', '3.1.1.3', 'α/β hydrolase', 320, 159.5, 14.0, 50.2, 'Multi-motif'],
    ['Glucose Oxidase', '1GAL', '1.1.3.4', 'FAD-binding', 581, 358.0, 34.0, 38.4, 'Single-motif'],
    ['Carbonic Anhydrase II', '1CA2', '4.2.1.1', 'All-β', 256, 186.0, 16.0, 27.4, 'Single-motif'],
    ['Tropinone Reductase-II', '2AE2', '1.1.1.184', 'Rossmann SDR', 259, 190.0, 15.0, 26.7, 'Single-motif'],
    ['Subtilisin BPN\'', '1SBT', '3.4.21.62', 'α/β sandwich', 275, 210.0, 16.0, 23.5, 'Single-motif'],
    ['Lysozyme', '1LSE', '3.2.1.17', 'α+β', 129, 103.0, 7.0, 20.4, 'Single-motif'],
    ['TEM-1 β-Lactamase', '1BTL', '3.5.2.6', 'α/β DD-peptidase', 263, 231.0, 14.0, 12.0, 'Single-motif'],
    ['Triosephosphate Isomerase', '1TIM', '5.3.1.1', 'TIM barrel', 247, 232.0, 15.0, 6.0, 'Single-motif'],
    ['P. cepacia Lipase (old)', '3LIP', '3.1.1.3', 'α/β hydrolase', 320, 384.1, 17.0, -20.4, 'Single-motif (deprecated)'],
]
for i, row in enumerate(sre_data, 4):
    for j, val in enumerate(row):
        ws2.cell(row=i, column=j+1, value=val)
        style_cell(ws2, i, j+1)
    # Color SRE column
    sre = row[7]
    sre_cell = ws2.cell(row=i, column=8)
    if sre > 30:
        sre_cell.fill = best_fill
    elif sre < 0:
        sre_cell.fill = bad_fill
    elif sre < 15:
        sre_cell.fill = warn_fill

# Bar chart
chart = BarChart()
chart.type = "col"
chart.title = "Size Reduction Efficiency (%)"
chart.y_axis.title = "SRE (%)"
chart.x_axis.title = "Enzyme"
chart.style = 10
chart.width = 22
chart.height = 14

data_ref = Reference(ws2, min_col=8, min_row=3, max_row=11, max_col=8)
cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=11)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "8B4513"
ws2.add_chart(chart, "A14")

auto_width(ws2)
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['D'].width = 22

# ============================================================
# SHEET 3: Active Site Geometry
# ============================================================
ws3 = wb.create_sheet("2-Active Site Geometry")

ws3.merge_cells('A1:H1')
ws3['A1'] = 'Active Site Geometry Preservation — Catalytic Cα RMSD Analysis'
ws3['A1'].font = Font(name='Calibri', bold=True, size=13, color='8B4513')

geo_headers = ['Enzyme', 'Cat. Residue Count', 'Cat. Residues', 'Mean RMSD (Å)', 'SD RMSD (Å)',
               'RMSD Range (Å)', '<0.5 Å (%)', '<1.0 Å (%)']
for j, h in enumerate(geo_headers, 1):
    ws3.cell(row=3, column=j, value=h)
style_header_row(ws3, 3, len(geo_headers))

geo_data = [
    ['TEM-1 β-Lactamase', 4, 'S70,K73,S130,E166', 0.13, 0.04, '0.06–0.28', 100.0, 100.0],
    ['Carbonic Anhydrase II', 5, 'H94,H96,H119,T199,E106', 0.19, 0.05, '0.10–0.37', 100.0, 100.0],
    ['Triosephosphate Isomerase', 2, 'H95,E165', 0.33, 0.05, '0.18–0.45', 100.0, 100.0],
    ['P. cepacia Lipase (new)', 3, 'S87,D264,H286', 0.44, 0.15, '0.17–0.78', 72.0, 100.0],
    ['Subtilisin BPN\'', 3, 'D32,H64,S221', 0.50, 0.06, '0.33–0.64', 48.0, 100.0],
    ['Glucose Oxidase', 2, 'H516,H559', 0.56, 0.07, '0.36–0.81', 20.0, 100.0],
    ['Tropinone Reductase-II', 2, 'Y155,Y159', 0.62, 0.05, '0.48–0.77', 2.0, 100.0],
    ['Lysozyme', 2, 'E35,D52', 0.68, 0.08, '0.49–0.89', 2.0, 100.0],
]
for i, row in enumerate(geo_data, 4):
    for j, val in enumerate(row):
        ws3.cell(row=i, column=j+1, value=val)
        style_cell(ws3, i, j+1)
    # Color <0.5Å column
    pct = row[6]
    pct_cell = ws3.cell(row=i, column=7)
    if pct == 100.0:
        pct_cell.fill = best_fill
    elif pct < 20:
        pct_cell.fill = warn_fill

auto_width(ws3)
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['F'].width = 16

# ============================================================
# SHEET 4: Sequence Design Quality
# ============================================================
ws4 = wb.create_sheet("3-Sequence Design (MPNN)")

ws4.merge_cells('A1:H1')
ws4['A1'] = 'ProteinMPNN Sequence Design Quality'
ws4['A1'].font = Font(name='Calibri', bold=True, size=13, color='8B4513')

mpnn_headers = ['Enzyme', 'Design Chain(s)', 'N Sequences', 'MPNN Best', 'MPNN Mean', 
                'MPNN SD', 'MPNN Worst', 'Seq Recovery']
for j, h in enumerate(mpnn_headers, 1):
    ws4.cell(row=3, column=j, value=h)
style_header_row(ws4, 3, len(mpnn_headers))

mpnn_data = [
    ['Triosephosphate Isomerase', 'B', 400, 0.862, 1.031, 0.061, 1.233, '~5%'],
    ['Tropinone Reductase-II', 'B', 400, 0.889, 1.080, 0.064, 1.275, '~4%'],
    ['Glucose Oxidase', 'B', 400, 0.897, 1.045, 0.053, 1.232, '~4%'],
    ['Carbonic Anhydrase II', 'B', 400, 0.897, 1.105, 0.068, 1.349, '~5%'],
    ['P. cepacia Lipase (new)', 'D', 1600, 0.901, 1.129, 0.080, 1.445, '~11%'],
    ['Subtilisin BPN\'', 'B', 400, 0.927, 1.092, 0.066, 1.260, '~4%'],
    ['TEM-1 β-Lactamase', 'B', 400, 0.956, 1.103, 0.056, 1.235, '~4%'],
    ['Lysozyme', 'B', 400, 1.001, 1.206, 0.100, 1.449, '~5%'],
    ['P. cepacia Lipase (old)', 'B', 400, 0.914, 1.064, 0.054, 1.238, '~3%'],
]
for i, row in enumerate(mpnn_data, 4):
    for j, val in enumerate(row):
        ws4.cell(row=i, column=j+1, value=val)
        style_cell(ws4, i, j+1)
    # Color best score
    best = row[3]
    if isinstance(best, (int, float)):
        best_cell = ws4.cell(row=i, column=4)
        if best < 0.9:
            best_cell.fill = best_fill

auto_width(ws4)
ws4.column_dimensions['A'].width = 28

# ============================================================
# SHEET 5: ESM-2 Embedding Analysis
# ============================================================
ws5 = wb.create_sheet("4-ESM2 Embeddings")

ws5.merge_cells('A1:G1')
ws5['A1'] = 'ESM-2 Embedding Analysis — Three-Way Comparison'
ws5['A1'].font = Font(name='Calibri', bold=True, size=13, color='8B4513')

esm_headers = ['Enzyme', 'Sim_intra (Design-Design)', 'Sim_native (Native-Design)', 
               'Random Baseline', 'Δ (Intra − Random)', 'Δ (Native − Random)', 'Interpretation']
for j, h in enumerate(esm_headers, 1):
    ws5.cell(row=3, column=j, value=h)
style_header_row(ws5, 3, len(esm_headers))

esm_data = [
    ['Lysozyme', 0.917, 0.706, 0.967, -0.050, -0.261, 'Most divergent from native'],
    ['Subtilisin BPN\'', 0.936, 0.814, 0.967, -0.031, -0.153, 'Moderate native divergence'],
    ['TEM-1 β-Lactamase', 0.940, 0.759, 0.967, -0.027, -0.208, 'Strong native divergence'],
    ['Triosephosphate Isomerase', 0.945, 0.835, 0.967, -0.022, -0.132, 'Closest to native sequence'],
    ['Glucose Oxidase', 0.946, 0.843, 0.967, -0.021, -0.124, 'Closest to native (large fixed motif)'],
    ['Carbonic Anhydrase II', 0.925, 0.786, 0.967, -0.042, -0.181, 'Diverse designs, native divergence'],
    ['Tropinone Reductase-II', 0.937, 0.810, 0.967, -0.030, -0.157, 'Moderate native divergence'],
    ['P. cepacia Lipase (new)', 0.908, 'N/A', 0.967, -0.059, 'N/A', 'Most diverse designs (multi-motif)'],
]
for i, row in enumerate(esm_data, 4):
    for j, val in enumerate(row):
        ws5.cell(row=i, column=j+1, value=val)
        style_cell(ws5, i, j+1)

# Add note about three-tier ordering
note_row = len(esm_data) + 5
ws5.merge_cells(f'A{note_row}:G{note_row}')
ws5[f'A{note_row}'] = 'Interpretation: Three-tier ordering: Random (0.967) > Design-Design (0.91–0.95) > Native-Design (0.71–0.84). Designed sequences occupy a distinct embedding-space region—more diverse than random, significantly divergent from native wild-type.'
ws5[f'A{note_row}'].font = Font(name='Calibri', italic=True, size=10, color='666666')

auto_width(ws5)
ws5.column_dimensions['A'].width = 28
ws5.column_dimensions['G'].width = 35

# ============================================================
# SHEET 6: Failure Mode Analysis
# ============================================================
ws6 = wb.create_sheet("5-Failure Modes")

ws6.merge_cells('A1:F1')
ws6['A1'] = 'Systematic Failure Mode Classification'
ws6['A1'].font = Font(name='Calibri', bold=True, size=13, color='8B4513')

fm_headers = ['Type', 'Description', 'Affected Enzymes', 'SRE Impact', 'Root Cause', 'Mitigation Strategy']
for j, h in enumerate(fm_headers, 1):
    ws6.cell(row=3, column=j, value=h)
style_header_row(ws6, 3, len(fm_headers))

fm_data = [
    ['Type A', 'Insufficient Compression', 'TIM (6%), TEM-1 (12%)', 'Severe',
     'Active site loop architecture integral to substrate binding; RFdiffusion cannot compress without distorting motif',
     'Consider loop grafting, domain-level redesign, or testing shorter target length ranges'],
    ['Type B', 'Contig Strategy Dependence', 'pc_lipase (old: −20%)', 'Critical',
     'Dispersed catalytic residues (200 aa span) forced oversized fixed region in single-motif mode (184 aa fixed)',
     'Multi-motif scaffolding (3 independent short segments, 15 aa total fixed) → SRE improved to +50%'],
    ['Type C', 'Cofactor Pocket Distortion Risk', 'Glucose Ox (38%), Tropinone Red (28%)', 'Moderate',
     'Backbone generation did not explicitly model FAD/NADP+ cofactor binding; designed pockets may be distorted',
     'Use RFdiffusion3 for all-atom cofactor modeling or post-design molecular docking validation'],
]
for i, row in enumerate(fm_data, 4):
    for j, val in enumerate(row):
        ws6.cell(row=i, column=j+1, value=val)
        style_cell(ws6, i, j+1)
    # Color type cell
    type_cell = ws6.cell(row=i, column=1)
    if 'Type A' in str(row[0]):
        type_cell.fill = bad_fill
    elif 'Type B' in str(row[0]):
        type_cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    elif 'Type C' in str(row[0]):
        type_cell.fill = warn_fill

auto_width(ws6)
ws6.column_dimensions['A'].width = 12
ws6.column_dimensions['B'].width = 30
ws6.column_dimensions['D'].width = 16
ws6.column_dimensions['E'].width = 55
ws6.column_dimensions['F'].width = 55

# ============================================================
# SHEET 7: Contig Design Reference
# ============================================================
ws7 = wb.create_sheet("6-Contig Configurations")

ws7.merge_cells('A1:G1')
ws7['A1'] = 'RFdiffusion Contig Configuration Reference'
ws7['A1'].font = Font(name='Calibri', bold=True, size=13, color='8B4513')

contig_headers = ['Enzyme', 'PDB', 'Native (aa)', 'Fixed Motif', 'Contig String', 
                  'Target Length', 'Design Chain(s)']
for j, h in enumerate(contig_headers, 1):
    ws7.cell(row=3, column=j, value=h)
style_header_row(ws7, 3, len(contig_headers))

contig_data = [
    ['Lysozyme', '1LSE', 129, 'A35–A52 (18 aa)', '[A35-52/0 72-98]', '72–98', 'B'],
    ['Subtilisin BPN\'', '1SBT', 275, 'A30–A65 (36 aa)', '[A30-65/0 150-200]', '150–200', 'B'],
    ['TEM-1 β-Lactamase', '1BTL', 263, 'A68–A132 (65 aa)', '[A68-132/0 140-190]', '140–190', 'B'],
    ['Triosephosphate Isomerase', '1TIM', 247, 'A93–A167 (75 aa)', '[A93-167/0 130-180]', '130–180', 'B'],
    ['Glucose Oxidase', '1GAL', 581, 'A514–A561 (48 aa)', '[A514-561/0 250-350]', '250–350', 'B'],
    ['Carbonic Anhydrase II', '1CA2', 256, 'A92–A121 (30 aa)', '[A92-121/0 130-180]', '130–180', 'B'],
    ['Tropinone Reductase-II', '2AE2', 259, 'A137–A161 (25 aa)', '[A137-161/0 140-190]', '140–190', 'B'],
    ['P. cepacia Lipase (v1)', '3LIP', 320, 'A85–A268 (184 aa)', '[A85-268/0 170-230]', '170–230', 'B'],
    ['P. cepacia Lipase (v2)', '3LIP', 320, 'A85-89,A262-266,A284-288 (15 aa)', '[A85-89/0 5-15/0 A262-266/0 5-15/0 A284-288/0 100-150]', '5-15 link + 100-150 scaffold', 'D (B/C motifs)'],
]
for i, row in enumerate(contig_data, 4):
    for j, val in enumerate(row):
        ws7.cell(row=i, column=j+1, value=val)
        style_cell(ws7, i, j+1)

auto_width(ws7)
ws7.column_dimensions['A'].width = 28
ws7.column_dimensions['E'].width = 55
ws7.column_dimensions['F'].width = 28
ws7.column_dimensions['G'].width = 22

# Freeze panes on all sheets
for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
    ws.freeze_panes = ws.cell(row=4, column=1)

# Save
wb.save(OUT_PATH)
print("Excel workbook saved: {}".format(OUT_PATH))
print("Sheets: {}".format(wb.sheetnames))
